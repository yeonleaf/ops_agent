#!/usr/bin/env python3
"""
이메일 첨부파일 처리 모듈
- 첨부파일 다운로드 및 저장
- LLM을 이용한 파일 내용 분석
- Vector DB에 첨부파일 내용 임베딩
- 첨부파일 검색 및 조회
"""

import os
import base64
import hashlib
import mimetypes
from pathlib import Path
from typing import Dict, List, Optional, Any, Tuple
from dataclasses import dataclass
from datetime import datetime

# 파일 처리 라이브러리
import PyPDF2
from PIL import Image
import docx
import openpyxl
import json

# LLM 및 임베딩
from langchain_openai import AzureChatOpenAI
from langchain_core.messages import HumanMessage

# 로깅
from module.logging_config import get_logger

# Vector DB
from chromadb_singleton import ChromaDBSingleton
from vector_db_models import VectorDBManager


@dataclass
class AttachmentMetadata:
    """첨부파일 메타데이터"""
    file_id: str
    original_filename: str
    file_path: str
    file_size: int
    mime_type: str
    file_hash: str
    created_at: str
    analysis_summary: Optional[str] = None
    extracted_text: Optional[str] = None
    keywords: List[str] = None
    file_type_category: Optional[str] = None  # 'image', 'document', 'spreadsheet', 'pdf', 'other'


@dataclass
class ProcessedAttachment:
    """처리된 첨부파일 정보"""
    metadata: AttachmentMetadata
    vector_db_ids: List[str]  # Vector DB에 저장된 청크들의 ID
    analysis_result: Dict[str, Any]


class AttachmentProcessor:
    """첨부파일 처리 클래스"""

    def __init__(self, storage_dir: str = "attachments", chunk_size: int = 1000):
        self.logger = get_logger(__name__)
        self.storage_dir = Path(storage_dir)
        self.storage_dir.mkdir(exist_ok=True)
        self.chunk_size = chunk_size

        # LLM 클라이언트 초기화
        self.llm_client = None
        self._init_llm_client()

        # Vector DB 매니저
        self.vector_db = VectorDBManager()

        # 지원하는 파일 유형
        self.supported_types = {
            'application/pdf': 'pdf',
            'application/vnd.openxmlformats-officedocument.wordprocessingml.document': 'document',
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet': 'spreadsheet',
            'text/plain': 'text',
            'text/csv': 'text',
            'image/png': 'image',
            'image/jpeg': 'image',
            'image/jpg': 'image',
            'image/gif': 'image',
            'application/json': 'text'
        }

    def _init_llm_client(self):
        """LLM 클라이언트 초기화"""
        try:
            self.llm_client = AzureChatOpenAI(
                azure_endpoint=os.getenv("AZURE_OPENAI_ENDPOINT"),
                api_key=os.getenv("AZURE_OPENAI_API_KEY"),
                api_version=os.getenv("AZURE_OPENAI_API_VERSION"),
                azure_deployment=os.getenv("AZURE_OPENAI_DEPLOYMENT_NAME"),
                temperature=0.1
            )
            self.logger.info("LLM 클라이언트 초기화 성공")
        except Exception as e:
            self.logger.error(f"LLM 클라이언트 초기화 실패: {e}")

    def save_attachment(self, file_data: bytes, filename: str, mime_type: str) -> AttachmentMetadata:
        """첨부파일을 로컬 저장소에 저장"""
        try:
            # 파일 해시 생성 (중복 방지)
            file_hash = hashlib.sha256(file_data).hexdigest()

            # 파일 확장자 추출
            _, ext = os.path.splitext(filename)
            if not ext and mime_type:
                ext = mimetypes.guess_extension(mime_type) or ''

            # 저장할 파일명 생성
            safe_filename = f"{file_hash[:16]}_{filename}"
            file_path = self.storage_dir / safe_filename

            # 파일이 이미 존재하면 기존 메타데이터 반환
            if file_path.exists():
                self.logger.info(f"파일이 이미 존재함: {safe_filename}")
                return AttachmentMetadata(
                    file_id=file_hash,
                    original_filename=filename,
                    file_path=str(file_path),
                    file_size=len(file_data),
                    mime_type=mime_type,
                    file_hash=file_hash,
                    created_at=datetime.now().isoformat()
                )

            # 파일 저장
            with open(file_path, 'wb') as f:
                f.write(file_data)

            self.logger.info(f"첨부파일 저장 완료: {safe_filename}")

            return AttachmentMetadata(
                file_id=file_hash,
                original_filename=filename,
                file_path=str(file_path),
                file_size=len(file_data),
                mime_type=mime_type,
                file_hash=file_hash,
                created_at=datetime.now().isoformat()
            )

        except Exception as e:
            self.logger.error(f"첨부파일 저장 실패: {e}")
            raise

    def extract_text_content(self, metadata: AttachmentMetadata) -> str:
        """파일에서 텍스트 내용 추출"""
        try:
            file_path = Path(metadata.file_path)

            if not file_path.exists():
                self.logger.warning(f"파일이 존재하지 않음: {file_path}")
                return ""

            mime_type = metadata.mime_type

            if mime_type == 'application/pdf':
                return self._extract_pdf_text(file_path)
            elif mime_type == 'application/vnd.openxmlformats-officedocument.wordprocessingml.document':
                return self._extract_docx_text(file_path)
            elif mime_type == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet':
                return self._extract_xlsx_text(file_path)
            elif mime_type.startswith('text/'):
                return self._extract_plain_text(file_path)
            elif mime_type.startswith('image/'):
                return self._extract_image_text(file_path)
            else:
                self.logger.info(f"지원하지 않는 파일 형식: {mime_type}")
                return ""

        except Exception as e:
            self.logger.error(f"텍스트 추출 실패: {e}")
            return ""

    def _extract_pdf_text(self, file_path: Path) -> str:
        """PDF에서 텍스트 추출"""
        try:
            text = ""
            with open(file_path, 'rb') as file:
                reader = PyPDF2.PdfReader(file)
                for page in reader.pages:
                    text += page.extract_text() + "\n"
            return text.strip()
        except Exception as e:
            self.logger.error(f"PDF 텍스트 추출 실패: {e}")
            return ""

    def _extract_docx_text(self, file_path: Path) -> str:
        """Word 문서에서 텍스트 추출"""
        try:
            doc = docx.Document(file_path)
            text = ""
            for paragraph in doc.paragraphs:
                text += paragraph.text + "\n"
            return text.strip()
        except Exception as e:
            self.logger.error(f"DOCX 텍스트 추출 실패: {e}")
            return ""

    def _extract_xlsx_text(self, file_path: Path) -> str:
        """Excel 파일에서 텍스트 추출"""
        try:
            workbook = openpyxl.load_workbook(file_path)
            text = ""
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                text += f"[Sheet: {sheet_name}]\n"
                for row in sheet.iter_rows(values_only=True):
                    row_text = "\t".join(str(cell) if cell is not None else "" for cell in row)
                    if row_text.strip():
                        text += row_text + "\n"
                text += "\n"
            return text.strip()
        except Exception as e:
            self.logger.error(f"XLSX 텍스트 추출 실패: {e}")
            return ""

    def _extract_plain_text(self, file_path: Path) -> str:
        """텍스트 파일 읽기"""
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                return f.read()
        except UnicodeDecodeError:
            try:
                with open(file_path, 'r', encoding='euc-kr') as f:
                    return f.read()
            except Exception as e:
                self.logger.error(f"텍스트 파일 읽기 실패: {e}")
                return ""

    def _extract_image_text(self, file_path: Path) -> str:
        """이미지에서 OCR로 텍스트 추출"""
        try:
            # 1. Azure Vision API 사용 (우선순위 1)
            azure_text = self._extract_image_text_azure(file_path)
            if azure_text and len(azure_text.strip()) > 10:  # 의미있는 텍스트가 추출된 경우
                self.logger.info(f"Azure Vision으로 이미지 텍스트 추출 성공: {file_path}")
                return azure_text

            # 2. pytesseract 사용 (폴백)
            tesseract_text = self._extract_image_text_tesseract(file_path)
            if tesseract_text and len(tesseract_text.strip()) > 5:
                self.logger.info(f"Tesseract로 이미지 텍스트 추출 성공: {file_path}")
                return tesseract_text

            # 3. 텍스트 추출 실패 시 이미지 파일명과 기본 정보
            image_info = self._get_basic_image_info(file_path)
            self.logger.warning(f"이미지 텍스트 추출 실패, 기본 정보만 반환: {file_path}")
            return image_info

        except Exception as e:
            self.logger.error(f"이미지 텍스트 추출 중 오류: {e}")
            return f"[Image: {file_path.name} - 텍스트 추출 실패]"

    def _extract_image_text_azure(self, file_path: Path) -> str:
        """Azure Computer Vision API를 사용한 OCR"""
        try:
            import requests

            endpoint = os.getenv("AZURE_VISION_ENDPOINT")
            key = os.getenv("AZURE_VISION_KEY")

            if not endpoint or not key:
                self.logger.warning("Azure Vision 설정이 없습니다.")
                return ""

            # 이미지 파일 읽기
            with open(file_path, 'rb') as image_data:
                image_bytes = image_data.read()

            # OCR API 호출
            ocr_url = f"{endpoint}/vision/v3.2/read/analyze"
            headers = {
                'Ocp-Apim-Subscription-Key': key,
                'Content-Type': 'application/octet-stream'
            }

            response = requests.post(ocr_url, headers=headers, data=image_bytes)
            response.raise_for_status()

            # 결과 URL 가져오기
            operation_location = response.headers["Operation-Location"]

            # 결과 대기 및 가져오기 (최대 30초)
            import time
            for i in range(15):  # 2초씩 15번 = 30초
                time.sleep(2)
                result_response = requests.get(operation_location, headers={'Ocp-Apim-Subscription-Key': key})
                result = result_response.json()

                if result["status"] == "succeeded":
                    # 텍스트 추출
                    extracted_text = ""
                    for page in result.get("analyzeResult", {}).get("readResults", []):
                        for line in page.get("lines", []):
                            extracted_text += line.get("text", "") + "\n"

                    return extracted_text.strip()
                elif result["status"] == "failed":
                    self.logger.error("Azure Vision OCR 실패")
                    break

            self.logger.warning("Azure Vision OCR 시간 초과")
            return ""

        except Exception as e:
            self.logger.warning(f"Azure Vision OCR 오류: {e}")
            return ""

    def _extract_image_text_tesseract(self, file_path: Path) -> str:
        """Tesseract를 사용한 OCR (폴백)"""
        try:
            import pytesseract
            from PIL import Image

            # 이미지 열기
            image = Image.open(file_path)

            # OCR 수행 (한국어 + 영어)
            text = pytesseract.image_to_string(image, lang='kor+eng')

            return text.strip()

        except ImportError:
            self.logger.warning("pytesseract가 설치되지 않았습니다.")
            return ""
        except Exception as e:
            self.logger.warning(f"Tesseract OCR 오류: {e}")
            return ""

    def _get_basic_image_info(self, file_path: Path) -> str:
        """기본 이미지 정보 추출"""
        try:
            from PIL import Image

            with Image.open(file_path) as img:
                width, height = img.size
                format_name = img.format
                mode = img.mode

                # 기본 이미지 정보
                info_text = f"""
이미지 파일: {file_path.name}
크기: {width}x{height}
형식: {format_name}
모드: {mode}
파일 크기: {file_path.stat().st_size} bytes

[이 이미지에서 텍스트를 추출할 수 없었습니다.
이미지 내용이 중요한 경우 수동으로 확인해주세요.]
                """.strip()

                return info_text

        except Exception as e:
            return f"[Image: {file_path.name} - 정보 추출 실패: {str(e)}]"

    def analyze_with_llm(self, metadata: AttachmentMetadata, content: str) -> Dict[str, Any]:
        """LLM을 사용하여 파일 내용 분석"""
        try:
            if not self.llm_client or not content.strip():
                return {"error": "LLM 클라이언트 없음 또는 내용 없음"}

            # 내용이 너무 길면 앞부분만 사용
            if len(content) > 4000:
                content = content[:4000] + "... (truncated)"

            analysis_prompt = f"""
다음은 업무 관련 이메일의 첨부파일 내용입니다. 이 파일을 분석하여 다음 정보를 JSON 형태로 제공해주세요:

파일명: {metadata.original_filename}
파일 타입: {metadata.mime_type}

파일 내용:
{content}

분석 요청사항:
1. summary: 파일 내용의 간단한 요약 (2-3문장)
2. keywords: 주요 키워드 추출 (5-10개)
3. category: 파일 카테고리 ("문서", "보고서", "스프레드시트", "이미지", "기타")
4. business_relevance: 업무 관련성 ("높음", "보통", "낮음")
5. key_points: 주요 내용 포인트 (3-5개)

JSON 형태로만 응답해주세요.
"""

            response = self.llm_client.invoke([HumanMessage(content=analysis_prompt)])

            try:
                # JSON 응답 파싱
                analysis_result = json.loads(response.content)
                self.logger.info(f"LLM 분석 완료: {metadata.original_filename}")
                return analysis_result

            except json.JSONDecodeError:
                self.logger.warning(f"LLM 응답 JSON 파싱 실패: {response.content}")
                return {
                    "summary": response.content[:200],
                    "keywords": [],
                    "category": "기타",
                    "business_relevance": "보통",
                    "key_points": []
                }

        except Exception as e:
            self.logger.error(f"LLM 분석 실패: {e}")
            return {"error": str(e)}

    def create_chunks(self, content: str, metadata: AttachmentMetadata) -> List[Dict[str, Any]]:
        """파일 내용을 청크로 분할"""
        if not content.strip():
            return []

        chunks = []
        words = content.split()

        for i in range(0, len(words), self.chunk_size):
            chunk_words = words[i:i + self.chunk_size]
            chunk_text = " ".join(chunk_words)

            chunk_metadata = {
                "file_id": metadata.file_id,
                "filename": metadata.original_filename,
                "mime_type": metadata.mime_type,
                "chunk_index": len(chunks),
                "file_size": metadata.file_size,
                "source": "attachment"
            }

            chunks.append({
                "text": chunk_text,
                "metadata": chunk_metadata
            })

        return chunks

    def store_in_vector_db(self, chunks: List[Dict[str, Any]], ticket_id: str,
                         analysis_result: Dict[str, Any]) -> List[str]:
        """Vector DB에 첨부파일 청크 저장"""
        try:
            from vector_db_models import AttachmentChunk
            vector_db_ids = []

            for chunk in chunks:
                chunk_metadata = chunk["metadata"]

                # AttachmentChunk 객체 생성
                attachment_chunk = AttachmentChunk(
                    chunk_id=f"attachment_{ticket_id}_{chunk_metadata['file_id']}_{chunk_metadata['chunk_index']}",
                    ticket_id=ticket_id,
                    file_id=chunk_metadata["file_id"],
                    original_filename=chunk_metadata["filename"],
                    mime_type=chunk_metadata["mime_type"],
                    chunk_index=chunk_metadata["chunk_index"],
                    content=chunk["text"],
                    file_size=chunk_metadata["file_size"],
                    analysis_summary=analysis_result.get("summary"),
                    keywords=analysis_result.get("keywords", []),
                    file_category=analysis_result.get("category"),
                    business_relevance=analysis_result.get("business_relevance"),
                    created_at=datetime.now().isoformat(),
                    source="attachment"
                )

                # Vector DB에 저장
                success = self.vector_db.add_attachment_chunk(attachment_chunk)

                if success:
                    vector_db_ids.append(attachment_chunk.chunk_id)
                    self.logger.debug(f"첨부파일 청크 저장 성공: {attachment_chunk.chunk_id}")
                else:
                    self.logger.error(f"첨부파일 청크 저장 실패: {attachment_chunk.chunk_id}")

            self.logger.info(f"Vector DB에 {len(vector_db_ids)}개 첨부파일 청크 저장 완료")
            return vector_db_ids

        except Exception as e:
            self.logger.error(f"Vector DB 저장 실패: {e}")
            return []

    def process_attachment_from_base64(self,
                                     base64_data: str,
                                     filename: str,
                                     mime_type: str,
                                     ticket_id: str) -> Optional[ProcessedAttachment]:
        """Base64 데이터에서 첨부파일 처리"""
        try:
            # Base64 디코드
            file_data = base64.b64decode(base64_data)

            # 파일 저장
            metadata = self.save_attachment(file_data, filename, mime_type)

            # 텍스트 추출
            extracted_text = self.extract_text_content(metadata)
            metadata.extracted_text = extracted_text

            # LLM 분석
            analysis_result = self.analyze_with_llm(metadata, extracted_text)
            if "summary" in analysis_result:
                metadata.analysis_summary = analysis_result["summary"]
                metadata.keywords = analysis_result.get("keywords", [])
                metadata.file_type_category = analysis_result.get("category", "기타")

            # 청크 생성
            chunks = self.create_chunks(extracted_text, metadata)

            # Vector DB 저장
            vector_db_ids = []
            if chunks:
                vector_db_ids = self.store_in_vector_db(chunks, ticket_id, analysis_result)

            self.logger.info(f"첨부파일 처리 완료: {filename}")

            return ProcessedAttachment(
                metadata=metadata,
                vector_db_ids=vector_db_ids,
                analysis_result=analysis_result
            )

        except Exception as e:
            self.logger.error(f"첨부파일 처리 실패: {e}")
            return None

    def search_attachments_in_vector_db(self, query: str, ticket_id: Optional[str] = None) -> List[Dict[str, Any]]:
        """Vector DB에서 첨부파일 검색"""
        try:
            # Vector DB에서 검색 (content_type이 attachment인 것만)
            results = []

            # 검색 쿼리 실행 (Vector DB 검색 로직 구현 필요)
            # 현재는 기본 구조만 제공

            self.logger.info(f"첨부파일 검색 완료: {len(results)}개 결과")
            return results

        except Exception as e:
            self.logger.error(f"첨부파일 검색 실패: {e}")
            return []


# 사용 예제
if __name__ == "__main__":
    processor = AttachmentProcessor()

    # 테스트용 Base64 데이터로 첨부파일 처리
    # test_data = "SGVsbG8gV29ybGQhIFRoaXMgaXMgYSB0ZXN0IGZpbGUu"  # "Hello World! This is a test file."
    # result = processor.process_attachment_from_base64(
    #     test_data, "test.txt", "text/plain", "test_ticket_001"
    # )
    # print(f"처리 결과: {result}")