"""
리팩토링된 파일 처리 시스템
전략 패턴, 의존성 주입, 표준화된 에러 처리, 개선된 임시 파일 관리 적용
"""

import os
import json
import tempfile
import logging
from pathlib import Path
from typing import List, Dict, Any, Optional
from datetime import datetime
from dotenv import load_dotenv

from module.exceptions import (
    ProcessingError, 
    FileTypeNotSupportedError, 
    ContentExtractionError
)
from module.processors import TextBasedProcessor, LayoutBasedProcessor
from module.converters import ConverterFactory

# .env 파일 로드
load_dotenv()

# 로깅 설정
logger = logging.getLogger(__name__)


class FileTypeDetector:
    """파일 타입 판별기"""
    
    @staticmethod
    def detect_file_type(file_path: str) -> str:
        """파일 확장자로부터 문서 타입 판별"""
        ext = Path(file_path).suffix.lower()
        if ext in ['.docx', '.pptx', '.pdf', '.xlsx', '.txt', '.md', '.csv', '.xml']:
            return ext[1:]  # .docx -> docx
        else:
            raise FileTypeNotSupportedError(file_path, ext)
    
    @staticmethod
    def detect_content_type(file_path: str, doc_type: str) -> str:
        """콘텐츠가 text-based인지 layout-based인지 판별"""
        try:
            if doc_type == 'pdf':
                return FileTypeDetector._analyze_pdf_content(file_path)
            elif doc_type == 'docx':
                return FileTypeDetector._analyze_docx_content(file_path)
            elif doc_type == 'pptx':
                return FileTypeDetector._analyze_pptx_content(file_path)
            elif doc_type == 'xlsx':
                return FileTypeDetector._analyze_xlsx_content(file_path)
            elif doc_type == 'xml':
                return 'text_based'  # XML은 항상 text-based로 처리
            else:
                return 'text_based'
        except Exception as e:
            logger.warning(f"콘텐츠 타입 판별 오류: {e}")
            return 'layout_based'  # 오류 시 안전하게 layout-based로 분류
    
    @staticmethod
    def _analyze_pdf_content(file_path: str) -> str:
        """PDF 콘텐츠 분석"""
        try:
            import fitz
            doc = fitz.open(file_path)
            text_content = ""
            image_count = 0
            
            for page in doc:
                text_content += page.get_text()
                image_list = page.get_images()
                image_count += len(image_list)
            
            doc.close()
            
            # 텍스트가 충분하고 이미지가 적으면 text-based
            if len(text_content.strip()) > 100 and image_count < 3:
                return 'text_based'
            else:
                return 'layout_based'
                
        except Exception:
            return 'layout_based'
    
    @staticmethod
    def _analyze_docx_content(file_path: str) -> str:
        """DOCX 콘텐츠 분석"""
        try:
            from docx import Document
            doc = Document(file_path)
            text_content = ""
            image_count = 0
            
            for para in doc.paragraphs:
                text_content += para.text + "\n"
            
            # 이미지 개수 확인 (간단한 방법)
            for rel in doc.part.rels.values():
                if "image" in rel.target_ref:
                    image_count += 1
            
            if len(text_content.strip()) > 100 and image_count < 3:
                return 'text_based'
            else:
                return 'layout_based'
                
        except Exception:
            return 'layout_based'
    
    @staticmethod
    def _analyze_pptx_content(file_path: str) -> str:
        """
        PPTX 콘텐츠 분석 (파일 단위 - 슬라이드별 동적 분석으로 대체됨)
        
        참고: 이 함수는 더 이상 사용되지 않으며, 
        TextBasedProcessor._analyze_slide_content()에서 슬라이드별로 동적 분석을 수행합니다.
        """
        # PPTX는 슬라이드별로 동적 분석하므로 기본값 반환
        # 실제 분석은 TextBasedProcessor._analyze_slide_content()에서 수행
        return 'mixed_content'  # 혼합 콘텐츠로 분류
    
    @staticmethod
    def _analyze_xlsx_content(file_path: str) -> str:
        """XLSX 콘텐츠 분석"""
        try:
            from openpyxl import load_workbook
            wb = load_workbook(file_path, data_only=True)
            text_content = ""
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                for row in ws.iter_rows(values_only=True):
                    for cell in row:
                        if cell:
                            text_content += str(cell) + " "
            
            wb.close()
            
            # 엑셀은 대부분 text-based
            if len(text_content.strip()) > 50:
                return 'text_based'
            else:
                return 'layout_based'
                
        except Exception:
            return 'layout_based'


class FileProcessor:
    """리팩토링된 메인 파일 처리기"""
    
    def __init__(self, azure_processor=None, converter=None):
        """
        파일 처리기 초기화
        
        Args:
            azure_processor: Azure OpenAI 프로세서 (선택사항)
            converter: 파일 변환기 (선택사항)
        """
        self.azure_processor = azure_processor
        self.converter = converter
        
        # 프로세서 인스턴스 생성
        self.text_processor = TextBasedProcessor(azure_processor)
        self.layout_processor = LayoutBasedProcessor(azure_processor, converter)
        
        # 처리 통계
        self.processing_stats = {
            "total_files": 0,
            "successful_files": 0,
            "failed_files": 0,
            "total_chunks": 0
        }
    
    def process_file(self, file_path: str) -> Dict[str, Any]:
        """
        파일을 처리하고 결과를 반환
        
        Args:
            file_path: 처리할 파일 경로
            
        Returns:
            처리 결과 딕셔너리
        """
        try:
            logger.info(f"파일 처리 시작: {file_path}")
            
            # 파일 존재 확인
            if not os.path.exists(file_path):
                raise ProcessingError(f"파일을 찾을 수 없습니다: {file_path}", file_path)
            
            # 1. 파일 타입 판별
            doc_type = FileTypeDetector.detect_file_type(file_path)
            content_type = FileTypeDetector.detect_content_type(file_path, doc_type)
            
            logger.info(f"파일 타입: {doc_type}, 콘텐츠 타입: {content_type}")
            
            # 2. 파일 타입별 직접 처리 (이중 하이브리드 vs 단순 변환)
            if doc_type in ['pptx', 'pdf', 'docx']:
                # 이중 하이브리드 방식: 요소 단위 + Vision 분석
                processing_method = "dual_path_hybrid"
                chunks = self.text_processor.process(file_path)
            elif doc_type in ['xlsx', 'txt', 'md', 'csv']:
                # 단순 변환 방식: 요소 단위 분석만
                processing_method = "simple_conversion"
                chunks = self.text_processor.process(file_path)
            else:
                # 기존 방식 (fallback)
                if content_type == 'text_based':
                    processor = self.text_processor
                    processing_method = "text_based"
                else:
                    processor = self.layout_processor
                    processing_method = "layout_based"
                chunks = processor.process(file_path)
            
            # 4. 결과 구성
            result = {
                "file_info": {
                    "file_path": file_path,
                    "file_name": Path(file_path).name,
                    "file_type": doc_type,
                    "content_type": content_type,
                    "processing_method": processing_method,
                    "total_chunks": len(chunks),
                    "processing_timestamp": datetime.now().isoformat()
                },
                "chunks": chunks,
                "processing_stats": {
                    "chunks_by_type": self._count_chunks_by_type(chunks),
                    "total_elements": len(chunks)
                }
            }
            
            # 5. 처리 통계 업데이트
            self._update_processing_stats(True, len(chunks))
            
            logger.info(f"파일 처리 완료: {file_path} -> {len(chunks)}개 청크")
            return result
            
        except Exception as e:
            logger.error(f"파일 처리 오류: {e}")
            self._update_processing_stats(False, 0)
            
            # 예외를 표준화된 에러 메시지로 변환
            if isinstance(e, ProcessingError):
                return e.to_dict()
            else:
                error = ProcessingError(str(e), file_path)
                return error.to_dict()
    
    def process_files_batch(self, file_paths: List[str]) -> List[Dict[str, Any]]:
        """
        여러 파일을 일괄 처리
        
        Args:
            file_paths: 처리할 파일 경로 리스트
            
        Returns:
            처리 결과 리스트
        """
        results = []
        
        for file_path in file_paths:
            try:
                result = self.process_file(file_path)
                results.append(result)
            except Exception as e:
                logger.error(f"일괄 처리 중 오류 ({file_path}): {e}")
                error_result = ProcessingError(str(e), file_path).to_dict()
                results.append(error_result)
        
        return results
    
    def _count_chunks_by_type(self, chunks: List[Dict[str, Any]]) -> Dict[str, int]:
        """청크 타입별 개수를 계산"""
        chunk_types = {}
        for chunk in chunks:
            element_type = chunk.get("metadata", {}).get("element_type", "unknown")
            chunk_types[element_type] = chunk_types.get(element_type, 0) + 1
        return chunk_types
    
    def _update_processing_stats(self, success: bool, chunk_count: int):
        """처리 통계 업데이트"""
        self.processing_stats["total_files"] += 1
        if success:
            self.processing_stats["successful_files"] += 1
            self.processing_stats["total_chunks"] += chunk_count
        else:
            self.processing_stats["failed_files"] += 1
    
    def get_processing_stats(self) -> Dict[str, Any]:
        """처리 통계 반환"""
        return self.processing_stats.copy()
    
    def reset_stats(self):
        """처리 통계 초기화"""
        self.processing_stats = {
            "total_files": 0,
            "successful_files": 0,
            "failed_files": 0,
            "total_chunks": 0
        }
    
    def save_result_to_file(self, result: Dict[str, Any], output_path: str, format: str = "json"):
        """
        처리 결과를 파일로 저장
        
        Args:
            result: 저장할 결과
            output_path: 출력 파일 경로
            format: 출력 형식 ("json" 또는 "md")
        """
        try:
            os.makedirs(os.path.dirname(output_path), exist_ok=True)
            
            if format.lower() == "json":
                with open(output_path, 'w', encoding='utf-8') as f:
                    json.dump(result, f, ensure_ascii=False, indent=2)
                logger.info(f"결과를 JSON 파일로 저장: {output_path}")
                
            elif format.lower() == "md":
                markdown_content = self._result_to_markdown(result)
                with open(output_path, 'w', encoding='utf-8') as f:
                    f.write(markdown_content)
                logger.info(f"결과를 Markdown 파일로 저장: {output_path}")
                
            else:
                raise ValueError(f"지원하지 않는 출력 형식: {format}")
                
        except Exception as e:
            logger.error(f"결과 저장 오류: {e}")
            raise ProcessingError(f"결과 저장 실패: {str(e)}")
    
    def _result_to_markdown(self, result: Dict[str, Any]) -> str:
        """결과를 Markdown 형식으로 변환"""
        if "error" in result:
            return f"# 오류 발생\n\n```\n{result.get('message', '알 수 없는 오류')}\n```"
        
        md_content = []
        
        # 파일 정보
        file_info = result.get("file_info", {})
        md_content.append(f"# {file_info.get('file_name', 'Unknown File')} 처리 결과")
        md_content.append("")
        md_content.append("## 📋 파일 정보")
        md_content.append(f"- **파일명**: {file_info.get('file_name', 'N/A')}")
        md_content.append(f"- **파일 타입**: {file_info.get('file_type', 'N/A')}")
        md_content.append(f"- **콘텐츠 타입**: {file_info.get('content_type', 'N/A')}")
        md_content.append(f"- **처리 방법**: {file_info.get('processing_method', 'N/A')}")
        md_content.append(f"- **총 청크 수**: {file_info.get('total_chunks', 0)}")
        md_content.append(f"- **처리 시간**: {file_info.get('processing_timestamp', 'N/A')}")
        md_content.append("")
        
        # 처리 통계
        processing_stats = result.get("processing_stats", {})
        md_content.append("## 📊 처리 통계")
        md_content.append(f"- **총 요소 수**: {processing_stats.get('total_elements', 0)}")
        
        chunks_by_type = processing_stats.get("chunks_by_type", {})
        for chunk_type, count in chunks_by_type.items():
            md_content.append(f"- **{chunk_type}**: {count}")
        md_content.append("")
        
        # 청크별 내용
        chunks = result.get("chunks", [])
        for i, chunk in enumerate(chunks, 1):
            metadata = chunk.get("metadata", {})
            text_content = chunk.get("text_chunk_to_embed", "")
            
            md_content.append(f"## 📄 청크 {i}")
            md_content.append("")
            md_content.append(f"**타입**: {metadata.get('element_type', 'unknown')}")
            md_content.append(f"**섹션**: {metadata.get('section_title', 'N/A')}")
            md_content.append(f"**페이지**: {metadata.get('page_number', 'N/A')}")
            md_content.append("")
            md_content.append("### 내용")
            md_content.append("```")
            md_content.append(text_content)
            md_content.append("```")
            md_content.append("")
            md_content.append("---")
            md_content.append("")
        
        return "\n".join(md_content) 